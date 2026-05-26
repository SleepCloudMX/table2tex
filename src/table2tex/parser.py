import re
from pathlib import Path

from table2tex.model import TableData, ColumnMeta
from table2tex.utils import parse_cell_value, strip_tex_formatting, split_tex_cells


def parse_table(filepath: str) -> TableData:
    """Dispatch to the appropriate parser based on file extension."""
    ext = Path(filepath).suffix.lower()
    if ext == '.md':
        return _parse_markdown(filepath)
    elif ext in ('.xls', '.xlsx'):
        return _parse_excel(filepath)
    elif ext == '.tex':
        return _parse_tex(filepath)
    else:
        raise ValueError(f"Unsupported format: {ext}")


# ---------------------------------------------------------------------------
# Markdown
# ---------------------------------------------------------------------------

def _parse_markdown(filepath: str) -> TableData:
    text = Path(filepath).read_text(encoding='utf-8')
    lines = text.splitlines()

    # Locate separator row (|---|...|)
    sep_idx = None
    for i, line in enumerate(lines):
        if re.match(r'^\s*\|?\s*[-:]{3,}\s*(\|\s*[-:]{3,}\s*)*\|?\s*$', line):
            sep_idx = i
            break

    if sep_idx is None:
        raise ValueError("No markdown table separator row found (expected |---|...|)")

    header_lines = lines[:sep_idx]
    data_lines = lines[sep_idx + 1:]

    def _split_md_row(row: str) -> list[str]:
        # Strip leading/trailing | then split
        r = row.strip()
        if r.startswith('|'):
            r = r[1:]
        if r.endswith('|'):
            r = r[:-1]
        return [c.strip() for c in r.split('|')]

    headers = [_split_md_row(hl) for hl in header_lines if hl.strip()]
    data = [_split_md_row(dl) for dl in data_lines if dl.strip()]

    # Normalise column count
    max_cols = max((len(h) for h in headers), default=0)
    max_cols = max(max_cols, max((len(d) for d in data), default=0))
    headers = [_pad_row(h, max_cols) for h in headers]
    data = [_pad_row(d, max_cols) for d in data]

    columns = _build_column_meta(data, max_cols)
    return TableData(headers=headers, data=data, columns=columns, source='md')


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _parse_excel(filepath: str, sheet_name: str | None = None) -> TableData:
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Read all rows as strings
    raw_rows: list[list[str]] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        str_row = [_excel_cell_str(c) for c in row]
        if any(s for s in str_row):  # skip fully empty rows
            raw_rows.append(str_row)

    if not raw_rows:
        raise ValueError("Excel sheet is empty")

    # Handle merged cells: fill None cells that are in merged ranges
    _fill_merged(raw_rows, ws)

    # First row = header
    headers = [raw_rows[0]]
    data = raw_rows[1:]

    max_cols = max((len(h) for h in headers), default=0)
    max_cols = max(max_cols, max((len(d) for d in data), default=0))
    headers = [_pad_row(h, max_cols) for h in headers]
    data = [_pad_row(d, max_cols) for d in data]

    columns = _build_column_meta(data, max_cols)
    return TableData(headers=headers, data=data, columns=columns, source='xlsx')


def _excel_cell_str(value) -> str:
    if value is None:
        return ''
    if isinstance(value, (int, float)):
        # preserve original formatting as much as possible
        if isinstance(value, float) and value == int(value):
            return str(int(value))
        return str(value)
    return str(value).strip()


def _fill_merged(rows: list[list[str]], ws) -> None:
    """Fill cells that are hidden by merged ranges with the top-left cell's value."""
    for merged_range in ws.merged_cells.ranges:
        min_col = merged_range.min_col - 1  # 0-based
        max_col = merged_range.max_col - 1
        min_row = merged_range.min_row - 1
        max_row = merged_range.max_row - 1

        top_left = ''
        if min_row < len(rows) and min_col < len(rows[min_row]):
            top_left = rows[min_row][min_col]

        for r in range(min_row, min(max_row + 1, len(rows))):
            for c in range(min_col, min(max_col + 1, len(rows[r]) if r < len(rows) else 0)):
                if r == min_row and c == min_col:
                    continue
                while len(rows[r]) <= c:
                    rows[r].append('')
                if not rows[r][c]:
                    rows[r][c] = top_left


# ---------------------------------------------------------------------------
# TeX
# ---------------------------------------------------------------------------

_TABLE_ENV_RE = re.compile(
    r'\\begin\{(array|tabular)\}(?:\[[^\]]*\])?\{([^}]*)\}(.*?)\\end\{\1\}',
    re.DOTALL,
)

_HLINE_RE = re.compile(r'\s*\\hline\s*')


def _parse_tex(filepath: str) -> TableData:
    text = Path(filepath).read_text(encoding='utf-8')

    m = _TABLE_ENV_RE.search(text)
    if not m:
        raise ValueError("No tabular or array environment found in TeX file")

    env_type = m.group(1)
    col_spec = m.group(2)
    body = m.group(3)
    body_start = m.start(3)

    # Split body by \\ to get raw row fragments (may contain \hline)
    raw_fragments = _split_tex_body(body)

    # Clean each fragment: remove \hline, track positions
    cell_rows: list[list[str]] = []
    tex_row_strs: list[str] = []
    hlines: set[int] = set()

    for frag in raw_fragments:
        # Check if this fragment starts with \hline
        has_hline = bool(_HLINE_RE.match(frag))
        cleaned = _HLINE_RE.sub('', frag).strip()
        if not cleaned:
            # It was only an \hline, record it for the row before which it appears
            hlines.add(len(cell_rows))
            continue

        cells = split_tex_cells(cleaned)
        if cells:
            if has_hline:
                hlines.add(len(cell_rows))
            cell_rows.append(cells)
            tex_row_strs.append(cleaned)

    if not cell_rows:
        raise ValueError("No data rows found in TeX table body")

    # First row = header
    headers = [[strip_tex_formatting(c) for c in cell_rows[0]]]
    tex_headers = [cell_rows[0]]

    data = []
    tex_rows = []
    for row in cell_rows[1:]:
        stripped = [strip_tex_formatting(c) for c in row]
        data.append(stripped)
        tex_rows.append(row)

    max_cols = max((len(h) for h in headers), default=0)
    max_cols = max(max_cols, max((len(d) for d in data), default=0))
    headers = [_pad_row(h, max_cols) for h in headers]
    data = [_pad_row(d, max_cols) for d in data]
    tex_headers = [_pad_row(h, max_cols) for h in tex_headers]
    tex_rows = [_pad_row(r, max_cols) for r in tex_rows]

    columns = _build_column_meta(data, max_cols)

    # Adjust hlines for header offset: hlines track positions in tex_rows (which includes headers)
    # Already correct since hlines was computed against cell_rows

    # Check for trailing \hline after last row
    trailing_hline = bool(re.search(r'\\hline\s*$', body))

    return TableData(
        headers=headers,
        data=data,
        columns=columns,
        source='tex',
        tex_source=text,
        tex_body_start=body_start,
        tex_body_end=m.end(3),
        tex_rows=tex_headers + tex_rows,
        tex_hlines=hlines,
        tex_env_type=env_type,
        tex_col_spec=col_spec,
        tex_trailing_hline=trailing_hline,
    )


def _split_tex_body(body: str) -> list[str]:
    """Split TeX table body into row fragments by \\ .
    Each fragment is the text between two \\ markers."""
    parts = re.split(r'\\\\(?:\*|\[[^\]]*\])?', body)
    return [p for p in parts if p.strip()]


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _pad_row(row: list[str], n: int) -> list[str]:
    while len(row) < n:
        row.append('')
    return row


def _build_column_meta(data: list[list[str]], num_cols: int) -> list[ColumnMeta]:
    columns = []
    for ci in range(num_cols):
        all_numeric = True
        for row in data:
            cell = row[ci] if ci < len(row) else ''
            is_num, _, _ = parse_cell_value(cell)
            if not cell.strip():  # empty cell = missing value, skip
                continue
            if not is_num:
                all_numeric = False
                break
        columns.append(ColumnMeta(index=ci, is_numeric=all_numeric))
    return columns
